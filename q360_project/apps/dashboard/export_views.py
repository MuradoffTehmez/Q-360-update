"""
Analytics Dashboard Export Views.
Excel və PDF export funksiyaları.
"""
from django.contrib.auth.decorators import login_required
from django.http import HttpResponse
from django.utils import timezone
from datetime import timedelta
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.chart import BarChart, PieChart, Reference
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib import colors
from reportlab.lib.units import inch
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, PageBreak
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont
from io import BytesIO

from apps.evaluations.models import EvaluationResult, EvaluationCampaign
from apps.accounts.models import User
from apps.departments.models import Department
from apps.competencies.models import Competency


@login_required
def export_analytics_excel(request):
    """
    Export analytics dashboard data to Excel.
    """
    # Create workbook
    wb = Workbook()

    # Remove default sheet
    wb.remove(wb.active)

    # Create sheets
    _create_overview_sheet(wb)
    _create_evaluations_sheet(wb)
    _create_departments_sheet(wb)
    _create_competencies_sheet(wb)
    _create_trends_sheet(wb)

    # Prepare response
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="Q360_Analytics_{timezone.now().strftime("%Y%m%d")}.xlsx"'

    # Save workbook to response
    wb.save(response)

    return response


def _create_overview_sheet(wb):
    """Create overview sheet with key metrics."""
    ws = wb.create_sheet('Ümumi Baxış', 0)

    # Header styling
    header_fill = PatternFill(start_color='1F4E78', end_color='1F4E78', fill_type='solid')
    header_font = Font(color='FFFFFF', bold=True, size=12)

    # Title
    ws['A1'] = 'Q360 Analytics Dashboard'
    ws['A1'].font = Font(size=16, bold=True, color='1F4E78')
    ws.merge_cells('A1:D1')

    ws['A2'] = f"Hesabat Tarixi: {timezone.now().strftime('%d.%m.%Y %H:%M')}"
    ws.merge_cells('A2:D2')

    # Key metrics
    ws['A4'] = 'Əsas Göstəricilər'
    ws['A4'].font = Font(size=14, bold=True)

    # Headers
    headers = ['Metrik', 'Dəyər', 'Dəyişiklik', 'Status']
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=5, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')

    # Data
    metrics = [
        ('İstifadəçi Sayı', User.objects.filter(is_active=True).count(), '+5%', '✓'),
        ('Aktiv Kampaniyalar', EvaluationCampaign.objects.filter(status='active').count(), '+2', '✓'),
        ('Tamamlanmış Qiymətləndirmələr', EvaluationResult.objects.filter(status='completed').count(), '+15%', '✓'),
        ('Ortalama Skor', round(EvaluationResult.objects.filter(status='completed').aggregate(avg=Avg('overall_score'))['avg'] or 0, 2), '+0.3', '✓'),
    ]

    for row_idx, (metric, value, change, status) in enumerate(metrics, start=6):
        ws.cell(row=row_idx, column=1, value=metric)
        ws.cell(row=row_idx, column=2, value=value)
        ws.cell(row=row_idx, column=3, value=change)
        ws.cell(row=row_idx, column=4, value=status)

    # Column widths
    ws.column_dimensions['A'].width = 30
    ws.column_dimensions['B'].width = 15
    ws.column_dimensions['C'].width = 15
    ws.column_dimensions['D'].width = 10


def _create_evaluations_sheet(wb):
    """Create evaluations statistics sheet."""
    ws = wb.create_sheet('Qiymətləndirmələr')

    # Import here to avoid circular import
    from django.db.models import Avg

    # Header
    ws['A1'] = 'Qiymətləndirmə Statistikası'
    ws['A1'].font = Font(size=14, bold=True)

    # Get last 30 days results
    thirty_days_ago = timezone.now() - timedelta(days=30)
    results = EvaluationResult.objects.filter(
        calculated_at__gte=thirty_days_ago,
        status='completed'
    ).select_related('evaluatee', 'campaign')

    # Headers
    headers = ['ID', 'İstifadəçi', 'Kampaniya', 'Ümumi Skor', 'Özqiymətləndirmə', 'Rəhbər', 'Həmkar', 'Tarix']
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=3, column=col, value=header)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color='D9E1F2', end_color='D9E1F2', fill_type='solid')

    # Data
    for row_idx, result in enumerate(results[:100], start=4):
        ws.cell(row=row_idx, column=1, value=result.id)
        ws.cell(row=row_idx, column=2, value=result.evaluatee.get_full_name())
        ws.cell(row=row_idx, column=3, value=result.campaign.title if result.campaign else 'N/A')
        ws.cell(row=row_idx, column=4, value=round(result.overall_score, 2))
        ws.cell(row=row_idx, column=5, value=round(result.self_score or 0, 2))
        ws.cell(row=row_idx, column=6, value=round(result.supervisor_score or 0, 2))
        ws.cell(row=row_idx, column=7, value=round(result.peer_score or 0, 2))
        ws.cell(row=row_idx, column=8, value=result.calculated_at.strftime('%d.%m.%Y'))

    # Auto-size columns
    for col in range(1, 9):
        ws.column_dimensions[chr(64 + col)].width = 20


def _create_departments_sheet(wb):
    """Create departments performance sheet."""
    ws = wb.create_sheet('Şöbələr')

    from django.db.models import Avg, Count

    # Header
    ws['A1'] = 'Şöbə Performansı'
    ws['A1'].font = Font(size=14, bold=True)

    # Headers
    headers = ['Şöbə', 'İşçi Sayı', 'Ortalama Skor', 'Qiymətləndirmə Sayı', 'Status']
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=3, column=col, value=header)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color='E2EFDA', end_color='E2EFDA', fill_type='solid')

    # Get department stats
    departments = Department.objects.annotate(
        employee_count=Count('user'),
        avg_score=Avg('user__evaluation_results__overall_score')
    ).order_by('-avg_score')

    for row_idx, dept in enumerate(departments, start=4):
        ws.cell(row=row_idx, column=1, value=dept.name)
        ws.cell(row=row_idx, column=2, value=dept.employee_count)
        ws.cell(row=row_idx, column=3, value=round(dept.avg_score or 0, 2))
        ws.cell(row=row_idx, column=4, value=dept.user.filter(evaluation_results__status='completed').count())

        # Status based on avg score
        avg = dept.avg_score or 0
        status = '✓ Yüksək' if avg >= 4.0 else '~ Orta' if avg >= 3.0 else '✗ Aşağı'
        ws.cell(row=row_idx, column=5, value=status)

    # Column widths
    for col in range(1, 6):
        ws.column_dimensions[chr(64 + col)].width = 20


def _create_competencies_sheet(wb):
    """Create competencies analysis sheet."""
    ws = wb.create_sheet('Kompetensiyalar')

    # Header
    ws['A1'] = 'Kompetensiya Analizi'
    ws['A1'].font = Font(size=14, bold=True)

    # Headers
    headers = ['Kompetensiya', 'Kateqoriya', 'Ortalama Skor', 'Qiymətləndirmə Sayı', 'Trend']
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=3, column=col, value=header)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color='FFF2CC', end_color='FFF2CC', fill_type='solid')

    # Get competency data
    from apps.evaluations.models import Response
    from django.db.models import Avg, Count

    competencies = Competency.objects.annotate(
        avg_rating=Avg('question__responses__rating'),
        response_count=Count('question__responses')
    ).filter(response_count__gt=0).order_by('-avg_rating')

    for row_idx, comp in enumerate(competencies[:50], start=4):
        ws.cell(row=row_idx, column=1, value=comp.name)
        ws.cell(row=row_idx, column=2, value=comp.category.name if comp.category else 'N/A')
        ws.cell(row=row_idx, column=3, value=round(comp.avg_rating or 0, 2))
        ws.cell(row=row_idx, column=4, value=comp.response_count)
        ws.cell(row=row_idx, column=5, value='↗' if comp.avg_rating >= 4.0 else '→')

    # Column widths
    for col in range(1, 6):
        ws.column_dimensions[chr(64 + col)].width = 25


def _create_trends_sheet(wb):
    """Create trends analysis sheet."""
    ws = wb.create_sheet('Trendlər')

    # Header
    ws['A1'] = 'Performans Trendləri (Son 6 Ay)'
    ws['A1'].font = Font(size=14, bold=True)

    # Monthly data
    from django.db.models import Avg
    from django.db.models.functions import TruncMonth

    monthly_avg = EvaluationResult.objects.filter(
        status='completed',
        calculated_at__gte=timezone.now() - timedelta(days=180)
    ).annotate(
        month=TruncMonth('calculated_at')
    ).values('month').annotate(
        avg_score=Avg('overall_score'),
        count=Count('id')
    ).order_by('month')

    # Headers
    headers = ['Ay', 'Ortalama Skor', 'Qiymətləndirmə Sayı', 'Dəyişiklik']
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=3, column=col, value=header)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color='FCE4D6', end_color='FCE4D6', fill_type='solid')

    prev_score = None
    for row_idx, data in enumerate(monthly_avg, start=4):
        month_name = data['month'].strftime('%B %Y')
        avg_score = round(data['avg_score'], 2)
        count = data['count']

        ws.cell(row=row_idx, column=1, value=month_name)
        ws.cell(row=row_idx, column=2, value=avg_score)
        ws.cell(row=row_idx, column=3, value=count)

        if prev_score:
            change = avg_score - prev_score
            change_str = f"+{round(change, 2)}" if change > 0 else f"{round(change, 2)}"
            ws.cell(row=row_idx, column=4, value=change_str)
        else:
            ws.cell(row=row_idx, column=4, value='-')

        prev_score = avg_score

    # Column widths
    for col in range(1, 5):
        ws.column_dimensions[chr(64 + col)].width = 20


@login_required
def export_analytics_pdf(request):
    """
    Export analytics dashboard to PDF.
    """
    buffer = BytesIO()

    # Create PDF
    doc = SimpleDocTemplate(
        buffer,
        pagesize=landscape(A4),
        rightMargin=30,
        leftMargin=30,
        topMargin=50,
        bottomMargin=30
    )

    # Container for PDF elements
    elements = []

    # Styles
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontSize=24,
        textColor=colors.HexColor('#1F4E78'),
        spaceAfter=30,
        alignment=1  # Center
    )

    heading_style = ParagraphStyle(
        'CustomHeading',
        parent=styles['Heading2'],
        fontSize=16,
        textColor=colors.HexColor('#1F4E78'),
        spaceAfter=12
    )

    # Title
    title = Paragraph('Q360 Analytics Dashboard', title_style)
    elements.append(title)

    subtitle = Paragraph(
        f"Hesabat Tarixi: {timezone.now().strftime('%d %B %Y, %H:%M')}",
        styles['Normal']
    )
    elements.append(subtitle)
    elements.append(Spacer(1, 20))

    # Key Metrics Section
    elements.append(Paragraph('Əsas Göstəricilər', heading_style))

    from django.db.models import Avg

    metrics_data = [
        ['Metrik', 'Dəyər', 'Status'],
        ['İstifadəçi Sayı', str(User.objects.filter(is_active=True).count()), '✓'],
        ['Aktiv Kampaniyalar', str(EvaluationCampaign.objects.filter(status='active').count()), '✓'],
        ['Tamamlanmış Qiymətləndirmələr', str(EvaluationResult.objects.filter(status='completed').count()), '✓'],
        ['Ortalama Skor', str(round(EvaluationResult.objects.filter(status='completed').aggregate(avg=Avg('overall_score'))['avg'] or 0, 2)), '✓'],
    ]

    metrics_table = Table(metrics_data, colWidths=[300, 150, 100])
    metrics_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1F4E78')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 12),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
        ('GRID', (0, 0), (-1, -1), 1, colors.black)
    ]))

    elements.append(metrics_table)
    elements.append(Spacer(1, 30))

    # Department Performance
    elements.append(Paragraph('Şöbə Performansı', heading_style))

    from django.db.models import Count

    departments = Department.objects.annotate(
        employee_count=Count('user'),
        avg_score=Avg('user__evaluation_results__overall_score')
    ).order_by('-avg_score')[:10]

    dept_data = [['Şöbə', 'İşçi Sayı', 'Ortalama Skor']]
    for dept in departments:
        dept_data.append([
            dept.name,
            str(dept.employee_count),
            str(round(dept.avg_score or 0, 2))
        ])

    dept_table = Table(dept_data, colWidths=[300, 150, 150])
    dept_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#E2EFDA')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.black),
        ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 11),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('BACKGROUND', (0, 1), (-1, -1), colors.white),
        ('GRID', (0, 0), (-1, -1), 1, colors.grey)
    ]))

    elements.append(dept_table)

    # Build PDF
    doc.build(elements)

    # Prepare response
    buffer.seek(0)
    response = HttpResponse(buffer.read(), content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="Q360_Analytics_{timezone.now().strftime("%Y%m%d")}.pdf"'

    return response
