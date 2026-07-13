"""
Utility functions for report generation.
"""
from io import BytesIO
from django.conf import settings
from apps.evaluations.models import Response, QuestionCategory


def calculate_radar_data(evaluation_result):
    """Calculate radar chart data for an evaluation result."""
    from apps.evaluations.models import EvaluationAssignment

    assignments = EvaluationAssignment.objects.filter(
        campaign=evaluation_result.campaign,
        evaluatee=evaluation_result.evaluatee,
        status='completed'
    )

    categories = QuestionCategory.objects.filter(is_active=True)

    radar_data = {
        'categories': [],
        'self': [],
        'others': [],
        'average': []
    }

    for category in categories:
        radar_data['categories'].append(category.name)

        # Scores by relationship
        self_scores = []
        other_scores = []

        for assignment in assignments:
            responses = Response.objects.filter(
                assignment=assignment,
                question__category=category,
                score__isnull=False
            )

            for response in responses:
                if assignment.relationship == 'self':
                    self_scores.append(response.score)
                else:
                    other_scores.append(response.score)

        # Calculate averages
        self_avg = sum(self_scores) / len(self_scores) if self_scores else 0
        other_avg = sum(other_scores) / len(other_scores) if other_scores else 0
        overall_avg = (self_avg + other_avg) / 2 if (self_avg or other_avg) else 0

        radar_data['self'].append(round(self_avg, 2))
        radar_data['others'].append(round(other_avg, 2))
        radar_data['average'].append(round(overall_avg, 2))

    return radar_data


def generate_pdf_report(evaluation_result):
    """Generate PDF report for evaluation result."""
    try:
        from reportlab.lib.pagesizes import letter, A4
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
        from reportlab.lib.units import inch
        from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, PageBreak
        from reportlab.lib import colors
        from reportlab.lib.enums import TA_CENTER, TA_LEFT
    except ImportError:
        # Fallback if reportlab not installed
        return b"PDF generation requires reportlab package"

    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4)
    story = []

    styles = getSampleStyleSheet()

    # Title
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontSize=24,
        textColor=colors.HexColor('#0d6efd'),
        spaceAfter=30,
        alignment=TA_CENTER
    )

    story.append(Paragraph('360° Qiymətləndirmə Hesabatı', title_style))
    story.append(Spacer(1, 20))

    # Employee info
    info_data = [
        ['İşçi:', evaluation_result.evaluatee.get_full_name()],
        ['Kampaniya:', evaluation_result.campaign.title],
        ['Tarix:', evaluation_result.calculated_at.strftime('%d.%m.%Y')],
        ['Ümumi Bal:', f'{evaluation_result.overall_score:.2f}' if evaluation_result.overall_score else 'N/A'],
    ]

    info_table = Table(info_data, colWidths=[2*inch, 4*inch])
    info_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (0, -1), colors.grey),
        ('TEXTCOLOR', (0, 0), (0, -1), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
        ('FONTNAME', (0, 0), (-1, -1), 'Helvetica'),
        ('FONTSIZE', (0, 0), (-1, -1), 12),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 12),
        ('GRID', (0, 0), (-1, -1), 1, colors.black)
    ]))

    story.append(info_table)
    story.append(Spacer(1, 30))

    # Scores by relationship
    scores_data = [
        ['Qiymətləndirmə Növü', 'Ortalama Bal'],
        ['Özünüdəyərləndirmə', f'{evaluation_result.self_score:.2f}' if evaluation_result.self_score else '-'],
        ['Rəhbər', f'{evaluation_result.supervisor_score:.2f}' if evaluation_result.supervisor_score else '-'],
        ['Həmkar', f'{evaluation_result.peer_score:.2f}' if evaluation_result.peer_score else '-'],
        ['Tabelik', f'{evaluation_result.subordinate_score:.2f}' if evaluation_result.subordinate_score else '-'],
    ]

    scores_table = Table(scores_data, colWidths=[3*inch, 3*inch])
    scores_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#0d6efd')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 14),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
        ('GRID', (0, 0), (-1, -1), 1, colors.black)
    ]))

    story.append(Paragraph('Qiymətləndirmə Nəticələri', styles['Heading2']))
    story.append(Spacer(1, 12))
    story.append(scores_table)

    # Build PDF
    doc.build(story)
    pdf_content = buffer.getvalue()
    buffer.close()

    return pdf_content


def generate_excel_report(campaign):
    """Generate Excel report for campaign results."""
    try:
        import openpyxl
        from openpyxl.styles import Font, Alignment, PatternFill
    except ImportError:
        return b"Excel generation requires openpyxl package"

    from apps.evaluations.models import EvaluationResult

    # Create workbook
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Nəticələr"

    # Header style
    header_fill = PatternFill(start_color="0d6efd", end_color="0d6efd", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True, size=12)

    # Headers
    headers = [
        'İşçi ID', 'Ad Soyad', 'Şöbə', 'Vəzifə',
        'Ümumi Bal', 'Özüm', 'Rəhbər', 'Həmkar', 'Tabelik',
        'Qiymətləndirən Sayı', 'Tamamlanma %'
    ]

    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center')

    # Data
    results = EvaluationResult.objects.filter(
        campaign=campaign
    ).select_related('evaluatee', 'evaluatee__department').order_by('-overall_score')

    for row, result in enumerate(results, start=2):
        ws.cell(row=row, column=1).value = result.evaluatee.employee_id or result.evaluatee.username
        ws.cell(row=row, column=2).value = result.evaluatee.get_full_name()
        ws.cell(row=row, column=3).value = str(result.evaluatee.department) if result.evaluatee.department else '-'
        ws.cell(row=row, column=4).value = result.evaluatee.position or '-'
        ws.cell(row=row, column=5).value = float(result.overall_score) if result.overall_score else 0
        ws.cell(row=row, column=6).value = float(result.self_score) if result.self_score else 0
        ws.cell(row=row, column=7).value = float(result.supervisor_score) if result.supervisor_score else 0
        ws.cell(row=row, column=8).value = float(result.peer_score) if result.peer_score else 0
        ws.cell(row=row, column=9).value = float(result.subordinate_score) if result.subordinate_score else 0
        ws.cell(row=row, column=10).value = result.total_evaluators
        ws.cell(row=row, column=11).value = float(result.completion_rate)

    # Auto-size columns
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        adjusted_width = (max_length + 2)
        ws.column_dimensions[column_letter].width = adjusted_width

    # Save to BytesIO
    buffer = BytesIO()
    wb.save(buffer)
    excel_content = buffer.getvalue()
    buffer.close()

    return excel_content


def generate_csv_report(results_queryset, include_fields=None):
    """Generate CSV report from queryset."""
    import csv
    from io import StringIO

    output = StringIO()
    writer = csv.writer(output)

    # Default fields if not specified
    if not include_fields:
        include_fields = [
            'employee_id', 'full_name', 'department', 'position',
            'overall_score', 'self_score', 'supervisor_score',
            'peer_score', 'subordinate_score'
        ]

    # Headers
    headers = []
    for field in include_fields:
        headers.append(field.replace('_', ' ').title())
    writer.writerow(headers)

    # Data rows
    for result in results_queryset:
        row = []
        for field in include_fields:
            if field == 'employee_id':
                row.append(result.evaluatee.employee_id or result.evaluatee.username)
            elif field == 'full_name':
                row.append(result.evaluatee.get_full_name())
            elif field == 'department':
                row.append(str(result.evaluatee.department) if result.evaluatee.department else '-')
            elif field == 'position':
                row.append(result.evaluatee.position or '-')
            elif field == 'overall_score':
                row.append(float(result.overall_score) if result.overall_score else 0)
            elif field == 'self_score':
                row.append(float(result.self_score) if result.self_score else 0)
            elif field == 'supervisor_score':
                row.append(float(result.supervisor_score) if result.supervisor_score else 0)
            elif field == 'peer_score':
                row.append(float(result.peer_score) if result.peer_score else 0)
            elif field == 'subordinate_score':
                row.append(float(result.subordinate_score) if result.subordinate_score else 0)
            elif field == 'campaign':
                row.append(result.campaign.title if result.campaign else '-')
            elif field == 'calculated_at':
                row.append(result.calculated_at.strftime('%d.%m.%Y %H:%M') if result.calculated_at else '-')
            else:
                row.append('-')
        writer.writerow(row)

    csv_content = output.getvalue()
    output.close()

    return csv_content.encode('utf-8-sig')  # BOM for Excel compatibility


def build_dataset_excel(title, columns, rows, metadata=None):
    """Generate a generic Excel workbook from dataset."""
    try:
        import openpyxl
        from openpyxl.styles import Alignment, Font, PatternFill
    except ImportError as exc:
        raise RuntimeError('Excel export requires openpyxl') from exc

    metadata = metadata or {}
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = title[:31] if title else "Report"

    header_fill = PatternFill(start_color="0d6efd", end_color="0d6efd", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)

    for col_index, column in enumerate(columns, start=1):
        cell = ws.cell(row=1, column=col_index, value=column)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    for row_index, row in enumerate(rows, start=2):
        for col_index, value in enumerate(row, start=1):
            ws.cell(row=row_index, column=col_index).value = value

    for column_cells in ws.columns:
        max_length = max(len(str(cell.value)) if cell.value else 0 for cell in column_cells)
        ws.column_dimensions[column_cells[0].column_letter].width = max(max_length + 2, 12)

    if metadata:
        meta_sheet = wb.create_sheet("Metadata")
        meta_sheet["A1"].value = "Açar"
        meta_sheet["B1"].value = "Dəyər"
        for index, (key, value) in enumerate(metadata.items(), start=2):
            meta_sheet.cell(row=index, column=1, value=str(key))
            meta_sheet.cell(row=index, column=2, value=str(value))

    buffer = BytesIO()
    wb.save(buffer)
    return buffer.getvalue()


def build_dataset_pdf(title, columns, rows, metadata=None):
    """Generate a PDF table from dataset."""
    try:
        from reportlab.lib import colors
        from reportlab.lib.pagesizes import A4, landscape
        from reportlab.lib.styles import getSampleStyleSheet
        from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle
    except ImportError as exc:
        raise RuntimeError('PDF export requires reportlab') from exc

    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=landscape(A4))
    styles = getSampleStyleSheet()

    story = [Paragraph(title or "Hesabat", styles['Title']), Spacer(1, 18)]

    data = [columns] + rows
    table = Table(data, repeatRows=1)
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#0d6efd')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
    ]))
    story.append(table)

    if metadata:
        story.append(Spacer(1, 12))
        story.append(Paragraph("Məlumat xülasəsi", styles['Heading2']))
        for key, value in metadata.items():
            story.append(Paragraph(f"<b>{key}:</b> {value}", styles['BodyText']))

    doc.build(story)
    return buffer.getvalue()


def build_dataset_csv(title, columns, rows):
    """Generate CSV bytes from dataset."""
    import csv
    import io

    output = io.StringIO()
    writer = csv.writer(output)
    writer.writerow(columns)
    writer.writerows(rows)
    return output.getvalue().encode('utf-8-sig')
